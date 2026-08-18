"""
Local bulk resume parsing when external Bulk-Resume-Parser (port 8001) is unavailable.
Uses existing text_extraction + llm_service; stages uploads on disk; exports Excel.
Persists session/file progress to PostgreSQL (bulk_parse_sessions / bulk_parse_files).
Supports chunked uploads, ZIP extract, and parallel workers.
"""
from __future__ import annotations

import io
import os
import re
import threading
import time
import uuid
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any

from app.core import media_storage

# In-memory job store: job_id -> job dict
_local_jobs: dict[str, dict[str, Any]] = {}
_local_jobs_lock = threading.Lock()

_BULK_EXPORT_DIR = media_storage.bulk_exports_dir()
_BULK_UPLOAD_DIR = media_storage.bulk_uploads_dir()

ALLOWED_EXT = {'pdf', 'docx', 'png', 'jpg', 'jpeg', 'webp', 'tif', 'tiff'}
# Legacy .doc is not extractable by current text_extraction — reject before staging.
EXCEL_HEADERS = [
    'Filename',
    'Name',
    'Email',
    'Phone',
    'LinkedIn',
    'GitHub',
    'Current Location',
    'Preferred Location',
    'Summary',
    'Skills',
    'Experience',
    'Education',
    'Certifications',
    'Total Experience Years',
    'ParseStatus',
    'ParseNotes',
]

BULK_OCR_RETRY_DPI = 300
BULK_MIN_TEXT_CHARS = 30

# OCR/extract can run in parallel; LLM calls are gated by OLLAMA_MAX_CONCURRENT.
# Higher default now that most clean resumes skip Ollama via deterministic path.
BULK_PARSE_MAX_WORKERS = max(1, min(24, int(os.getenv('BULK_PARSE_MAX_WORKERS', '6'))))
BULK_LLM_ATTEMPTS = max(1, min(4, int(os.getenv('BULK_LLM_ATTEMPTS', '2'))))
BULK_EXCEL_CHECKPOINT_EVERY = max(5, int(os.getenv('BULK_EXCEL_CHECKPOINT_EVERY', '25')))
BULK_SKIP_LLM_WHEN_DETERMINISTIC = os.getenv(
    'BULK_SKIP_LLM_WHEN_DETERMINISTIC', 'true'
).lower() in ('1', 'true', 'yes')


def _is_retryable_llm_error(err: str) -> bool:
    low = (err or '').lower()
    return any(
        token in low
        for token in (
            'timeout',
            'timed out',
            'connection',
            'temporarily',
            'unavailable',
            'overloaded',
            'busy',
            '429',
            '500',
            '502',
            '503',
            '504',
            'provider not available',
            'no healthy',
            'read timed out',
            'connect',
        )
    )


def _experience_titles_implausible(toon: dict) -> bool:
    """True when any experience title fails the hardened job-title gate."""
    from app.ai.parser.enrichment.resume_text_inference import is_plausible_job_title

    exp = toon.get('experience') if isinstance(toon, dict) else None
    if not isinstance(exp, list):
        return False
    for e in exp:
        if not isinstance(e, dict):
            continue
        title = (e.get('title') or e.get('role') or '').strip()
        if title and not is_plausible_job_title(title):
            return True
    return False


def _ocr_experience_slice_mushy(raw_text: str) -> bool:
    """Cheap OCR-quality check on Experience section — refuse det-skip when mushy."""
    from app.ai.document_intelligence.coverage.resume_coverage import (
        _experience_section_text,
        has_experience_section_evidence,
    )

    if not has_experience_section_evidence(raw_text or ''):
        return False
    body = _experience_section_text(raw_text or '')
    if not body or len(body.strip()) < 40:
        return False
    sample = body[:1200]
    letters = sum(1 for c in sample if c.isalnum())
    if letters < 20:
        return True
    non_alnum_ratio = 1.0 - (letters / max(len(sample), 1))
    if non_alnum_ratio > 0.45:
        return True
    # Broken token density: long glued ALLCAPS / no-space runs
    glued = re.findall(r'\b[A-Za-z]{20,}\b', sample)
    if len(glued) >= 3:
        return True
    spaced_words = [w for w in re.findall(r'[A-Za-z]{2,}', sample)]
    if spaced_words:
        weird = sum(
            1
            for w in spaced_words
            if re.search(r'[a-z]{3,}[A-Z]{3,}', w)
            or (w.isupper() and len(w) > 12)
        )
        if weird >= 4:
            return True
    return False


def _call_llm_throttled(raw_text: str, doc_type: str = 'resume'):
    """Queue LLM calls so Ollama is not overwhelmed on large batches."""
    from app.ai.parser.engine.ollama_limit import ollama_slot
    from app.integrations.openai.llm_service import call_llm

    with ollama_slot():
        return call_llm(raw_text, doc_type)


def _export_path(job_id: str) -> Path:
    return _BULK_EXPORT_DIR / f'{job_id}.xlsx'


def _staging_dir(job_id: str) -> Path:
    return _BULK_UPLOAD_DIR / job_id


def _safe_filename(name: str) -> str:
    """Basename + secure characters; keep uniqueness via caller if needed."""
    from werkzeug.utils import secure_filename

    base = secure_filename(Path(name).name) or 'file'
    return base


def _unique_staged_name(job_id: str, filename: str) -> str:
    """Avoid collisions when the same basename appears in multiple batches/zip entries."""
    safe = _safe_filename(filename)
    dest_dir = _staging_dir(job_id)
    dest_dir.mkdir(parents=True, exist_ok=True)
    candidate = safe
    stem = Path(safe).stem
    suffix = Path(safe).suffix
    n = 1
    while (dest_dir / candidate).exists():
        candidate = f'{stem}_{n}{suffix}'
        n += 1
    return candidate


def _persist_excel(job_id: str, rows: list[dict], append: bool = False) -> None:
    try:
        _BULK_EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        path = _export_path(job_id)
        if append and path.is_file():
            existing = _read_excel_rows(path)
            # Prefer newer rows for same Filename
            by_name = {r.get('Filename'): r for r in existing if r.get('Filename')}
            for r in rows:
                by_name[r.get('Filename')] = r
            merged = list(by_name.values()) if by_name else (existing + rows)
            path.write_bytes(_build_excel_bytes(merged))
        else:
            path.write_bytes(_build_excel_bytes(rows))
    except Exception as e:
        print(f"[local_bulk_parser] persist excel failed: {e}")


def _read_excel_rows(path: Path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(h) if h is not None else '' for h in next(rows_iter)]
    except StopIteration:
        wb.close()
        return []
    out = []
    for vals in rows_iter:
        row = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            row[h] = vals[i] if i < len(vals) and vals[i] is not None else ''
        if any(str(v).strip() for v in row.values()):
            out.append(row)
    wb.close()
    return out


def _as_text(value: Any) -> str:
    """Coerce LLM TOON values to stripped text (ints/lists/None-safe)."""
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return str(value).strip()
    if isinstance(value, (int, float)):
        return str(value).strip()
    if isinstance(value, list):
        parts = [_as_text(item) for item in value]
        return ' '.join(p for p in parts if p)
    if isinstance(value, dict):
        for key in ('name', 'title', 'text', 'value', 'label', 'email', 'phone', 'skill'):
            if key in value and value[key] is not None:
                return _as_text(value[key])
        return ''
    return str(value).strip()


def _normalize_phone(phone: Any) -> str:
    s = _as_text(phone)
    if not s:
        return ''
    # Keep leading +, digits, spaces, dashes, parentheses
    cleaned = re.sub(r'[^\d+\-\s().]', '', s).strip()
    try:
        from app.ai.document_intelligence.validation import validate_phone

        ok, _ = validate_phone(cleaned)
        if not ok:
            return ''
    except Exception:
        pass
    return cleaned


def _normalize_email(email: Any) -> str:
    s = _as_text(email).lower()
    if not s:
        return ''
    if '@' not in s:
        return s
    return s


def _cert_to_str(c: Any) -> str:
    if isinstance(c, dict):
        name = _as_text(c.get('name'))
        issuer = _as_text(c.get('issuer'))
        if name and issuer:
            return f'{name} ({issuer})'
        return name or issuer or _as_text(c)
    return _as_text(c)


def _skill_to_str(s: Any) -> str:
    if isinstance(s, dict):
        return _as_text(s.get('name') or s.get('skill') or s)
    return _as_text(s)


def _flatten_toon(toon: dict, filename: str, form: Any = None) -> dict:
    """Flatten one TOON resume to a row dict for Excel; prefer Form DTO for location/education."""
    if not isinstance(toon, dict):
        toon = {}
    person = toon.get('person') or {}
    if not isinstance(person, dict):
        person = {}
    skills = toon.get('skills') or []
    exp = toon.get('experience') or []
    edu = toon.get('education') or []
    certs = toon.get('certifications') or []

    if not isinstance(skills, list):
        skills = [skills] if skills else []
    if not isinstance(exp, list):
        exp = []
    if not isinstance(edu, list):
        edu = []
    if not isinstance(certs, list):
        certs = [certs] if certs else []

    exp_parts = []
    for e in exp[:25]:
        if not isinstance(e, dict):
            continue
        title = _as_text(e.get('title') or e.get('role'))
        company = _as_text(e.get('company'))
        fr = _as_text(e.get('from') or e.get('start'))
        to = _as_text(e.get('to') or e.get('end'))
        if (e.get('is_current') or str(to).lower() in ('present', 'current', 'now')) and not to:
            to = 'Present'
        desc = _as_text(e.get('description'))
        # Skip geo-only / empty noise rows
        if not title and not company:
            continue
        if re.match(r'(?i)^(india|pune|mumbai|remote)$', title) and re.match(
            r'(?i)^(india|pune|mumbai|remote)$', company or 'x'
        ):
            continue
        if title and company:
            chunk = f'{title} at {company}'
        elif title:
            chunk = title
        else:
            chunk = company
        if fr or to:
            chunk = f'{chunk} ({fr}-{to})'.strip()
        if desc:
            chunk = f'{chunk}: {desc[:400]}'
        if chunk and chunk not in ('at ()', '()'):
            exp_parts.append(chunk)

    edu_parts = []
    # Prefer Form DTO education (includes mapper heals / grounding filters)
    form_edu = None
    if form is not None:
        form_edu = getattr(form, 'education', None)
        if form_edu is None and isinstance(form, dict):
            form_edu = form.get('education')
    if form_edu:
        for e in list(form_edu)[:15]:
            if hasattr(e, 'degree'):
                degree = _as_text(getattr(e, 'degree', ''))
                inst = _as_text(getattr(e, 'institution', ''))
                year = _as_text(getattr(e, 'endMonth', '') or getattr(e, 'startMonth', ''))
            elif isinstance(e, dict):
                degree = _as_text(e.get('degree'))
                inst = _as_text(e.get('institution'))
                year = _as_text(e.get('endMonth') or e.get('startMonth') or e.get('to'))
            else:
                continue
            chunk = degree
            if inst:
                chunk = f'{chunk} - {inst}' if chunk else inst
            if year:
                chunk = f'{chunk} [{year}]' if chunk else year
            if chunk:
                edu_parts.append(chunk)
    if not edu_parts:
        for e in edu[:15]:
            if not isinstance(e, dict):
                continue
            degree = _as_text(e.get('degree'))
            inst = _as_text(e.get('institution'))
            field = _as_text(e.get('field'))
            year = _as_text(e.get('year') or e.get('to') or e.get('from'))
            chunk = degree
            if field:
                chunk = f'{chunk} ({field})' if chunk else field
            if inst:
                chunk = f'{chunk} - {inst}' if chunk else inst
            if year:
                chunk = f'{chunk} [{year}]' if chunk else year
            if chunk:
                edu_parts.append(chunk)

    skill_strs = [_skill_to_str(s) for s in skills[:50]]
    skill_strs = [s for s in skill_strs if s]
    cert_strs = [_cert_to_str(c) for c in certs[:30]]
    cert_strs = [c for c in cert_strs if c]

    years = toon.get('total_experience_years')
    if years is None or years == '':
        # Phase 6 safety net: recompute from experience dates / description ranges
        try:
            from app.ai.parser.enrichment.resume_text_inference import (
                compute_total_experience_years,
            )

            recomputed = compute_total_experience_years(
                [
                    {
                        'from': e.get('from') or e.get('start'),
                        'to': e.get('to') or e.get('end'),
                        'description': e.get('description') or '',
                        'role': e.get('title') or e.get('role') or '',
                        'company': e.get('company') or '',
                    }
                    for e in exp
                    if isinstance(e, dict)
                ]
            )
            years_out = recomputed if recomputed is not None else ''
        except Exception:
            years_out = ''
    else:
        years_out = years

    current_location = _as_text(
        person.get('location') or person.get('current_location')
    )
    preferred_location = _as_text(person.get('preferred_location')) or current_location
    if form is not None:
        form_cur = getattr(form, 'currentLocation', None)
        form_pref = getattr(form, 'preferredLocation', None)
        if form_cur is None and isinstance(form, dict):
            form_cur = form.get('currentLocation')
            form_pref = form.get('preferredLocation')
        if form_cur:
            current_location = _as_text(form_cur)
        if form_pref:
            preferred_location = _as_text(form_pref)
        elif current_location:
            preferred_location = current_location

    return {
        'Filename': filename,
        'Name': _as_text(person.get('name')),
        'Email': _normalize_email(person.get('email')),
        'Phone': _normalize_phone(person.get('phone')),
        'LinkedIn': _as_text(person.get('linkedin')),
        'GitHub': _as_text(person.get('github')),
        'Current Location': current_location,
        'Preferred Location': preferred_location,
        'Summary': _as_text(toon.get('summary'))[:2000],
        'Skills': ', '.join(skill_strs)[:4000],
        'Experience': '; '.join(exp_parts)[:8000],
        'Education': '; '.join(edu_parts)[:3000],
        'Certifications': ', '.join(cert_strs)[:2000],
        'Total Experience Years': years_out,
        'ParseStatus': 'ok',
        'ParseNotes': '',
    }


def _build_excel_bytes(rows: list[dict]) -> bytes:
    """Build .xlsx from list of row dicts using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Resumes'
    headers = EXCEL_HEADERS
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, '') for h in headers])
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(40, max(10, len(str(headers[col - 1])) + 2))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _bulk_timing_begin(filename: str, job_id: str | None = None):
    """Start a Developer Mode session for one bulk file (worker thread). Returns (ctx, wall_start) or (None, None)."""
    try:
        from app.core.developer_mode import is_developer_mode_enabled
        from app.core.request_context import start_request_context
        from app.core.timing_collector import timing_collector

        if not is_developer_mode_enabled():
            return None, None
        path = f"/api/admin/bulk-parse/{filename}"
        ctx = start_request_context(path=path, method="WORKER")
        if job_id:
            ctx.job_id = str(job_id)
        timing_collector.begin_session(
            request_id=ctx.request_id,
            started_at=ctx.started_at_iso,
            path=ctx.path,
            method=ctx.method,
            user_id=ctx.user_id,
            job_id=str(job_id) if job_id else None,
        )
        return ctx, time.perf_counter()
    except Exception:
        return None, None


def _bulk_timing_end(ctx, wall_start: float | None, *, failed: bool = False) -> None:
    try:
        if ctx is None:
            return
        from app.core.request_context import set_timing_context
        from app.core.timing_collector import timing_collector

        if failed:
            timing_collector.mark_error(ctx.request_id)
        wall_ms = None
        if wall_start is not None:
            wall_ms = (time.perf_counter() - wall_start) * 1000.0
        timing_collector.end_session(ctx.request_id, wall_duration_ms=wall_ms)
        set_timing_context(None)
    except Exception:
        pass


def _bulk_stage(stage: str, outcome: str, duration_ms: float = 0.0) -> None:
    try:
        from app.core.timing_collector import record_pipeline_stage

        record_pipeline_stage(
            stage,
            outcome,
            duration_ms=duration_ms,
            module="app.workers.bulk_parser",
        )
    except Exception:
        pass


_RESUME_CORE_COVERAGE_GAPS = frozenset(
    {'fullName', 'email', 'phone', 'location', 'education', 'experience'}
)
_COVERAGE_GAP_LABELS = {
    'fullName': 'name',
    'email': 'email',
    'phone': 'phone',
    'location': 'location',
    'education': 'education',
    'experience': 'experience',
}


def _coverage_gaps_from_form(form_dto) -> list[str]:
    """Named core coverage gaps from ApplicationFormDTO (JD-parity honesty)."""
    if form_dto is None:
        return []
    rows = getattr(form_dto, 'coverage', None)
    if rows is None and isinstance(form_dto, dict):
        rows = form_dto.get('coverage')
    if not isinstance(rows, list):
        return []
    gaps: list[str] = []
    for c in rows:
        if not isinstance(c, dict):
            continue
        field = str(c.get('field') or '')
        if field in _RESUME_CORE_COVERAGE_GAPS and c.get('status') == 'missing_with_evidence':
            gaps.append(field)
    return gaps


def _apply_coverage_parse_honesty(
    row: dict,
    form_dto,
    *,
    parse_status: str,
    note_bits: list[str],
) -> str:
    """Mark partial + ParseNotes when coverage still has named gaps with evidence."""
    gaps = _coverage_gaps_from_form(form_dto)
    if gaps:
        parse_status = 'partial'
        labels = [_COVERAGE_GAP_LABELS.get(g, g) for g in gaps]
        note_bits.append('coverage_gaps=' + ','.join(labels))
    row['ParseStatus'] = parse_status
    row['ParseNotes'] = '; '.join(note_bits)[:2000]
    return parse_status


def _process_one_file(args: tuple) -> tuple[str, dict | None, bool, str, str]:
    """
    Process a single staged file: extract → deterministic rules → LLM if needed.
    Returns (filename, row or None, failed: bool, message, code).
    Codes: insufficient_text | unsupported_format | llm | validation | empty_fields | exception | ok | partial
    """
    # args may be (filename, path) or (filename, path, job_id)
    if len(args) >= 3:
        filename, path, job_id = args[0], args[1], args[2]
    else:
        filename, path = args[0], args[1]
        job_id = None

    ctx, wall_start = _bulk_timing_begin(filename, job_id=job_id)
    failed = True
    try:
        result = _process_one_file_inner(filename, path, job_id=job_id)
        failed = bool(result[2])
        return result
    finally:
        _bulk_timing_end(ctx, wall_start, failed=failed)


def _maybe_enhance_layout(raw_text: str) -> tuple[str, str]:
    """Apply resume layout structuring when enabled. Returns (text, stage_outcome)."""
    try:
        from app.ai.parser.layout.detector import enhance_resume_text, is_layout_enabled

        if not is_layout_enabled():
            return raw_text, 'skipped'
        structured = enhance_resume_text(raw_text)
        if structured and len(structured.strip()) >= BULK_MIN_TEXT_CHARS:
            return structured, 'completed'
        return raw_text, 'completed'
    except Exception as e:
        print(f"[local_bulk_parser] layout enhance failed: {e}")
        return raw_text, 'failed'


def _persist_bulk_parse(
    *,
    job_id: str | None,
    filename: str,
    file_data: bytes,
    raw_text: str,
    toon: dict,
    confidence: float = 0.75,
) -> None:
    """
    Intentionally a no-op.

    Bulk must not write into raw_files/parsed_resumes used by single-file parse
    cache. Single-resume / apply cache is populated only by the single-parse path.
    Excel + bulk_parse_sessions remain the bulk outputs.
    """
    return


def _process_one_file_inner(
    filename: str,
    path: Path,
    *,
    job_id: str | None = None,
) -> tuple[str, dict | None, bool, str, str]:
    from app.ai.parser.text_extraction import extract_text
    from app.domains.recruitment.services.parsing_storage import validate_toon_format_bulk

    try:
        # Bulk text path: skip parse-cache lookup for throughput; layout when enabled
        _bulk_stage("cache", "skipped")

        data = path.read_bytes()
        last_extract_err = None
        raw_text = ""
        t_text = time.perf_counter()
        _IMAGE_EXTS = ('pdf', 'png', 'jpg', 'jpeg', 'webp', 'tif', 'tiff', 'bmp')
        ext = filename.lower().rsplit('.', 1)[-1] if '.' in filename else ''
        try:
            raw_text = extract_text(data, filename) or ""
        except Exception as extract_err:
            raw_text = ""
            last_extract_err = str(extract_err)[:200]

        if raw_text and '\x00' in raw_text:
            raw_text = raw_text.replace('\x00', '')

        def _looks_like_garbage_extract(s: str) -> bool:
            t = (s or '').strip()
            if len(t) < BULK_MIN_TEXT_CHARS:
                return True
            if len(t) > 400:
                return False
            alnum = sum(1 for c in t if c.isalnum())
            ratio = alnum / max(len(t), 1)
            has_token = bool(
                re.search(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', t)
                or re.search(r'\b[6-9]\d{9}\b|\+\d[\d\s\-()]{8,}\d', t)
                or re.search(r'(?i)\b(?:experience|education|skills|summary)\b', t)
            )
            return ratio < 0.35 and not has_token

        if (
            (len(raw_text.strip()) < BULK_MIN_TEXT_CHARS or _looks_like_garbage_extract(raw_text))
            and ext in _IMAGE_EXTS
        ):
            try:
                raw_text = extract_text(data, filename, dpi=BULK_OCR_RETRY_DPI) or ""
                if raw_text and '\x00' in raw_text:
                    raw_text = raw_text.replace('\x00', '')
                last_extract_err = None
            except Exception as retry_err:
                last_extract_err = str(retry_err)[:200]

        text_ms = (time.perf_counter() - t_text) * 1000.0
        if not raw_text or len(raw_text.strip()) < BULK_MIN_TEXT_CHARS:
            _bulk_stage("text", "failed", text_ms)
            _bulk_stage("persist_raw", "skipped")
            _bulk_stage("layout", "skipped")
            detail = last_extract_err or "insufficient text after OCR"
            low = (detail or "").lower()
            if "legacy .doc" in low or (
                filename.lower().endswith(".doc") and "not supported" in low
            ):
                return (
                    filename,
                    None,
                    True,
                    f"Skipped (unsupported format): {filename} - {detail}",
                    "unsupported_format",
                )
            return (
                filename,
                None,
                True,
                f"Skipped (insufficient text): {filename} - {detail}",
                "insufficient_text",
            )

        _bulk_stage("text", "completed", text_ms)

        t_layout = time.perf_counter()
        raw_text, layout_outcome = _maybe_enhance_layout(raw_text)
        _bulk_stage(
            "layout",
            layout_outcome,
            (time.perf_counter() - t_layout) * 1000.0,
        )

        # --- Intelligence Engine text path (shared with single-file parse) ---
        if BULK_SKIP_LLM_WHEN_DETERMINISTIC:
            try:
                from app.ai.parser.engine import parse_resume_text_via_engine
                from app.ai.parser.deterministic_resume import score_resume_toon

                from app.ai.document_intelligence.coverage.resume_coverage import (
                    has_experience_section_evidence,
                )

                det_toon, source, eng_notes, form_dto = parse_resume_text_via_engine(
                    raw_text,
                    allow_llm=False,
                    skip_llm_when_deterministic=True,
                )
                conf, missing, passes = score_resume_toon(
                    det_toon if isinstance(det_toon, dict) else {},
                    source_text=raw_text,
                )
                # Do not skip LLM when experience section exists but Excel would lack Experience
                exp_gap = (
                    has_experience_section_evidence(raw_text)
                    and isinstance(det_toon, dict)
                    and not (det_toon.get("experience") or [])
                )
                # Phase 6: refuse det-skip on garbled titles or OCR-mushy experience
                bad_titles = (
                    isinstance(det_toon, dict)
                    and _experience_titles_implausible(det_toon)
                )
                ocr_mush = _ocr_experience_slice_mushy(raw_text)
                if (
                    passes
                    and isinstance(det_toon, dict)
                    and not exp_gap
                    and not bad_titles
                    and not ocr_mush
                ):
                    t_val = time.perf_counter()
                    accept, notes, parse_status = validate_toon_format_bulk(det_toon, "resume")
                    _bulk_stage(
                        "validate",
                        "completed" if accept else "failed",
                        (time.perf_counter() - t_val) * 1000.0,
                    )
                    # Fast path: keep Ollama for true gaps (no experience / mushy OCR),
                    # not for ordinary "partial" validation notes.
                    if accept and parse_status in ("ok", "partial"):
                        t_persist = time.perf_counter()
                        row = _flatten_toon(det_toon, filename, form=form_dto)
                        note_bits = [f"source=engine:{source}", f"conf={conf:.2f}"]
                        note_bits.extend(eng_notes[:4])
                        if missing:
                            note_bits.append("weak=" + ",".join(missing[:6]))
                        if notes:
                            note_bits.append(notes)
                        parse_status = _apply_coverage_parse_honesty(
                            row,
                            form_dto,
                            parse_status=parse_status,
                            note_bits=note_bits,
                        )
                        _persist_bulk_parse(
                            job_id=job_id,
                            filename=filename,
                            file_data=data,
                            raw_text=raw_text,
                            toon=det_toon,
                            confidence=float(conf or 0.75),
                        )
                        _bulk_stage(
                            "persist_raw",
                            "skipped",
                            0.0,
                        )
                        _bulk_stage(
                            "persist",
                            "completed",
                            (time.perf_counter() - t_persist) * 1000.0,
                        )
                        if (
                            row.get("Name")
                            or row.get("Email")
                            or row.get("Phone")
                            or row.get("Skills")
                            or row.get("Experience")
                        ):
                            return (
                                filename,
                                row,
                                False,
                                f"Processing: {filename} (engine:{source})",
                                parse_status if parse_status in ("ok", "partial") else "ok",
                            )
            except Exception as det_err:
                print(f"[local_bulk_parser] engine det path failed for {filename}: {det_err}")

        # --- Slow path: engine with LLM (section-scoped + knowledge) ---
        toon = None
        last_err = None
        for attempt in range(BULK_LLM_ATTEMPTS):
            try:
                from app.ai.parser.engine import parse_resume_text_via_engine

                # Gate concurrent Ollama calls (OLLAMA_MAX_CONCURRENT)
                from app.ai.parser.engine.ollama_limit import ollama_slot

                with ollama_slot():
                    toon, source, eng_notes, form_dto = parse_resume_text_via_engine(
                        raw_text,
                        allow_llm=True,
                        skip_llm_when_deterministic=False,
                    )
                if not isinstance(toon, dict):
                    last_err = "Engine returned non-object"
                    toon = None
                    if attempt < BULK_LLM_ATTEMPTS - 1:
                        time.sleep(min(30, 2**attempt))
                    continue
                t_val = time.perf_counter()
                accept, notes, parse_status = validate_toon_format_bulk(toon, "resume")
                _bulk_stage(
                    "validate",
                    "completed" if accept else "failed",
                    (time.perf_counter() - t_val) * 1000.0,
                )
                if accept:
                    from app.ai.document_intelligence.coverage.resume_coverage import (
                        has_experience_section_evidence,
                    )

                    t_persist = time.perf_counter()
                    row = _flatten_toon(toon, filename, form=form_dto)
                    # Honest status: section evidence but still empty Experience → partial
                    if (
                        has_experience_section_evidence(raw_text)
                        and not (row.get("Experience") or "").strip()
                        and (
                            row.get("Name")
                            or row.get("Email")
                            or row.get("Phone")
                        )
                    ):
                        parse_status = "partial"
                    note_bits = [f"source=engine:{source}", f"status={parse_status}"]
                    note_bits.extend(eng_notes[:4])
                    if notes:
                        note_bits.append(notes)
                    parse_status = _apply_coverage_parse_honesty(
                        row,
                        form_dto,
                        parse_status=parse_status,
                        note_bits=note_bits,
                    )
                    _persist_bulk_parse(
                        job_id=job_id,
                        filename=filename,
                        file_data=data,
                        raw_text=raw_text,
                        toon=toon,
                        confidence=0.7,
                    )
                    _bulk_stage(
                        "persist_raw",
                        "skipped",
                        0.0,
                    )
                    _bulk_stage(
                        "persist",
                        "completed",
                        (time.perf_counter() - t_persist) * 1000.0,
                    )
                    if not (
                        row.get("Name")
                        or row.get("Email")
                        or row.get("Phone")
                        or row.get("Skills")
                        or row.get("Experience")
                    ):
                        last_err = "empty fields after flatten"
                        toon = None
                        continue
                    return (
                        filename,
                        row,
                        False,
                        f"Processing: {filename} (engine:{source})",
                        parse_status if parse_status in ("ok", "partial") else "ok",
                    )
                last_err = notes or "validation failed"
                toon = None
            except Exception as e:
                last_err = str(e)[:200]
                toon = None
                if _is_retryable_llm_error(last_err) and attempt < BULK_LLM_ATTEMPTS - 1:
                    time.sleep(min(45, 2**attempt + 1))
                    continue

        if last_err and "empty fields" in (last_err or ""):
            return (
                filename,
                None,
                True,
                f"Failed (empty fields): {filename}",
                "empty_fields",
            )
        code = "validation" if last_err and "validation" in (last_err or "").lower() else "llm"
        if last_err and any(
            x in (last_err or "").lower()
            for x in ("person.", "missing", "must not", "insufficient fields")
        ):
            code = "validation"
        return (
            filename,
            None,
            True,
            f"Failed (parse/validate): {filename} - {last_err}",
            code,
        )
    except Exception as e:
        return (filename, None, True, f"Failed: {filename} - {str(e)[:100]}", "exception")


def _load_staged_files(job_id: str) -> list[tuple[str, Path, str | None]]:
    """Return (filename, path, file_id) from durable DB + disk staging."""
    from app.domains.administration.repositories.bulk_session_db import list_queued_filenames, get_file_id_by_name

    names: list[str] = []
    with _local_jobs_lock:
        job = _local_jobs.get(job_id) or {}
        names = list(job.get('staged_filenames') or [])
    if not names:
        try:
            names = list_queued_filenames(job_id)
        except Exception:
            names = []
    out: list[tuple[str, Path, str | None]] = []
    for name in names:
        p = _staging_dir(job_id) / name
        if not p.is_file():
            continue
        file_id = None
        try:
            file_id = get_file_id_by_name(job_id, name)
        except Exception:
            file_id = None
        out.append((name, p, file_id))
    return out


def _worker(job_id: str, started_at: float, append: bool = False, worker_id: str | None = None) -> None:
    """Background: process staged files in parallel, update progress, write Excel."""
    from app.domains.administration.repositories.bulk_session_db import (
        claim_file_for_processing,
        finalize_session,
        heartbeat_session_lease,
        reclaim_stale_file_leases,
        update_file_status,
        update_session_progress,
    )

    wid = worker_id or f'pid-{os.getpid()}-{uuid.uuid4().hex[:8]}'
    try:
        reclaim_stale_file_leases()
    except Exception:
        pass

    with _local_jobs_lock:
        job = _local_jobs.get(job_id)
    if job and job.get('status') == 'cancelled':
        return
    # Ensure a local cache shell exists for progress UI even after worker restart.
    if not job:
        _ensure_job(job_id)
        with _local_jobs_lock:
            job = _local_jobs.get(job_id)
            if job:
                job['status'] = 'started'
                job['started_at'] = started_at

    files_list = _load_staged_files(job_id)
    total = len(files_list)
    results: list[dict] = []
    success_count = 0
    failed_count = 0
    max_workers = min(BULK_PARSE_MAX_WORKERS, max(1, total))

    def _run_one(item: tuple[str, Path, str | None]):
        filename, path, file_id = item
        if file_id:
            if not claim_file_for_processing(str(file_id), wid):
                return (filename, None, False, 'skipped-unclaimed', 'skipped', False)
        return (*_process_one_file((filename, path, job_id)), True)

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_file = {
            executor.submit(_run_one, item): item[0]
            for item in files_list
        }
        for future in as_completed(future_to_file):
            with _local_jobs_lock:
                j = _local_jobs.get(job_id)
            if j and j.get('status') == 'cancelled':
                break
            try:
                heartbeat_session_lease(job_id, wid)
            except Exception:
                pass
            filename, row, is_failed, message, code, claimed = future.result()
            if not claimed or code == 'skipped':
                continue
            if is_failed:
                failed_count += 1
                update_file_status(job_id, filename, 'Failed', error_message=message)
                with _local_jobs_lock:
                    if job_id in _local_jobs:
                        _local_jobs[job_id].setdefault('failed_filenames', []).append(filename)
                        _local_jobs[job_id].setdefault('failed_details', []).append(
                            {'filename': filename, 'error': message, 'code': code}
                        )
            else:
                if row:
                    results.append(row)
                success_count += 1
                notes = (row or {}).get('ParseNotes') or None
                update_file_status(job_id, filename, 'Completed', error_message=notes)
                with _local_jobs_lock:
                    if job_id in _local_jobs:
                        _local_jobs[job_id].setdefault('success_filenames', []).append(filename)
            done = success_count + failed_count
            update_session_progress(job_id, done, success_count, failed_count)
            with _local_jobs_lock:
                if job_id in _local_jobs:
                    _local_jobs[job_id]['processed_files'] = done
                    _local_jobs[job_id]['failed_files'] = failed_count
                    _local_jobs[job_id]['message'] = message
            # Checkpoint Excel during long runs so progress survives restarts
            if results and done > 0 and done % BULK_EXCEL_CHECKPOINT_EVERY == 0:
                _persist_excel(job_id, list(results), append=append)

    with _local_jobs_lock:
        if job_id in _local_jobs:
            _local_jobs[job_id]['status'] = 'completed'
            _local_jobs[job_id]['processed_files'] = success_count + failed_count
            _local_jobs[job_id]['failed_files'] = failed_count
            _local_jobs[job_id]['results'] = results
            _local_jobs[job_id]['message'] = f'Completed: {success_count} successful, {failed_count} failed'

    if results:
        _persist_excel(job_id, results, append=append)
    elif append and _export_path(job_id).is_file():
        pass  # keep existing workbook
    else:
        _persist_excel(job_id, [], append=False)

    finalize_session(job_id, started_at, success_count, failed_count)


def create_local_job(started_by=None, append: bool = False) -> tuple[str, dict]:
    """Create an empty bulk job ready to receive file chunks. Returns (job_id, payload)."""
    from app.domains.administration.repositories.bulk_session_db import create_empty_session

    job_id = None
    if started_by:
        job_id = create_empty_session(started_by)
    if not job_id:
        job_id = str(uuid.uuid4())

    _staging_dir(job_id).mkdir(parents=True, exist_ok=True)
    with _local_jobs_lock:
        _local_jobs[job_id] = {
            'status': 'pending',
            'total_files': 0,
            'processed_files': 0,
            'failed_files': 0,
            'results': [],
            'message': 'Waiting for uploads...',
            'failed_filenames': [],
            'success_filenames': [],
            'failed_details': [],
            'staged_filenames': [],
            'started_by': started_by,
            'append': bool(append),
            'started_at': None,
        }
    return job_id, {
        'job_id': job_id,
        'total_files': 0,
        'status': 'pending',
        'message': 'Bulk parse job created. Upload files, then start.',
    }


def _ensure_job(job_id: str, started_by=None, append: bool = False) -> dict | None:
    with _local_jobs_lock:
        job = _local_jobs.get(job_id)
        if job:
            if append:
                job['append'] = True
            return job
    # Recreate from DB owner if needed
    from app.domains.administration.repositories.bulk_session_db import get_session_owner

    owner = get_session_owner(job_id)
    if not owner and not started_by:
        return None
    _staging_dir(job_id).mkdir(parents=True, exist_ok=True)
    with _local_jobs_lock:
        if job_id not in _local_jobs:
            _local_jobs[job_id] = {
                'status': 'pending',
                'total_files': 0,
                'processed_files': 0,
                'failed_files': 0,
                'results': [],
                'message': 'Waiting for uploads...',
                'failed_filenames': [],
                'success_filenames': [],
                'failed_details': [],
                'staged_filenames': [],
                'started_by': owner or started_by,
                'append': bool(append),
                'started_at': None,
            }
        return _local_jobs[job_id]


def stage_files(job_id: str, files_list: list[tuple[str, bytes]], started_by=None) -> tuple[bool, dict]:
    """
    Stage resume bytes on disk for an existing (or newly ensured) job.
    files_list: list of (filename, bytes).
    """
    job = _ensure_job(job_id, started_by=started_by)
    if not job:
        return False, {'error': 'Job not found'}
    with _local_jobs_lock:
        status = job.get('status')
    if status not in ('pending', 'uploading'):
        return False, {'error': f'Cannot upload to job in status {status}'}

    from app.domains.administration.repositories.bulk_session_db import add_session_files, bump_session_total

    staged_names: list[str] = []
    file_bytes_for_db: list[bytes] = []
    for original_name, data in files_list:
        if not data:
            continue
        ext = original_name.rsplit('.', 1)[-1].lower() if '.' in original_name else ''
        if ext not in ALLOWED_EXT:
            continue
        staged_name = _unique_staged_name(job_id, original_name)
        (_staging_dir(job_id) / staged_name).write_bytes(data)
        staged_names.append(staged_name)
        file_bytes_for_db.append(data)

    if not staged_names:
        # Surface clear reject when only legacy .doc (or other junk) was uploaded
        rejected_doc = any(
            (n or '').lower().endswith('.doc') and not (n or '').lower().endswith('.docx')
            for n, _ in files_list
        )
        if rejected_doc:
            return False, {
                'error': (
                    'Unsupported format: legacy .doc is not accepted. '
                    'Convert to PDF or DOCX (or upload PNG/JPG for scanned resumes).'
                ),
                'code': 'unsupported_format',
            }
        return False, {'error': 'No valid resume files (PDF/DOCX/PNG/JPG/WEBP/TIFF)'}

    with _local_jobs_lock:
        if job_id in _local_jobs:
            _local_jobs[job_id]['status'] = 'uploading'
            _local_jobs[job_id].setdefault('staged_filenames', []).extend(staged_names)
            _local_jobs[job_id]['total_files'] = len(_local_jobs[job_id]['staged_filenames'])
            _local_jobs[job_id]['message'] = f"Uploaded {_local_jobs[job_id]['total_files']} file(s)"
            total = _local_jobs[job_id]['total_files']

    try:
        add_session_files(job_id, staged_names, file_bytes_for_db)
        bump_session_total(job_id, total)
    except Exception as exc:
        return False, {'error': f'Failed to catalog uploaded files: {exc}'}

    return True, {
        'job_id': job_id,
        'added_files': len(staged_names),
        'total_files': total,
        'status': 'uploading',
        'message': f'Added {len(staged_names)} file(s). Total staged: {total}.',
    }


def extract_zip_to_job(job_id: str, zip_bytes: bytes, started_by=None) -> tuple[bool, dict]:
    """Extract PDF/DOC/DOCX from a zip archive into the job staging directory."""
    job = _ensure_job(job_id, started_by=started_by)
    if not job:
        return False, {'error': 'Job not found'}
    with _local_jobs_lock:
        status = job.get('status')
    if status not in ('pending', 'uploading'):
        return False, {'error': f'Cannot upload to job in status {status}'}

    extracted: list[tuple[str, bytes]] = []
    try:
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
            for info in zf.infolist():
                if info.is_dir():
                    continue
                name = info.filename.replace('\\', '/')
                # Skip macOS junk / hidden
                parts = name.split('/')
                if any(p.startswith('.') or p == '__MACOSX' for p in parts):
                    continue
                base = parts[-1]
                if not base or '.' not in base:
                    continue
                ext = base.rsplit('.', 1)[-1].lower()
                if ext not in ALLOWED_EXT:
                    continue
                try:
                    data = zf.read(info)
                except Exception:
                    continue
                if data:
                    # Preserve nested path so subfolder duplicates stay unique
                    flat_name = '__'.join(p for p in parts if p) if len(parts) > 1 else base
                    extracted.append((flat_name, data))
    except zipfile.BadZipFile:
        return False, {'error': 'Invalid ZIP file'}
    except Exception as e:
        return False, {'error': f'ZIP extract failed: {e}'}

    if not extracted:
        return False, {'error': 'No valid resume files (PDF/DOCX/PNG/JPG/WEBP/TIFF) found in ZIP'}

    return stage_files(job_id, extracted, started_by=started_by)


def start_staged_job(job_id: str, append: bool | None = None) -> tuple[bool, dict]:
    """Start processing all staged files for a job (DB-backed, multi-worker safe)."""
    from app.domains.administration.repositories.bulk_session_db import (
        claim_session_lease,
        get_session_owner,
        get_session_progress,
        list_queued_filenames,
        reclaim_stale_file_leases,
    )

    try:
        reclaim_stale_file_leases()
    except Exception:
        pass

    # Rebuild local cache from durable state when this process did not handle upload.
    job = _ensure_job(job_id)
    db_names = []
    try:
        db_names = list_queued_filenames(job_id)
    except Exception:
        db_names = []

    with _local_jobs_lock:
        if not job:
            owner = None
            try:
                owner = get_session_owner(job_id)
            except Exception:
                owner = None
            if not owner and not db_names:
                progress = None
                try:
                    progress = get_session_progress(job_id)
                except Exception:
                    progress = None
                if not progress:
                    return False, {'error': 'Job not found'}
            job = _ensure_job(job_id, started_by=owner)
        if not job:
            return False, {'error': 'Job not found'}
        if job.get('status') == 'completed':
            return False, {'error': 'Job already completed'}
        # Prefer durable filename list from DB; fall back to memory / disk.
        names = list(job.get('staged_filenames') or [])
        if db_names:
            names = db_names
            job['staged_filenames'] = list(db_names)
        if not names:
            # Disk fallback for recovery when DB rows exist but names empty
            staging = _staging_dir(job_id)
            if staging.is_dir():
                names = sorted(p.name for p in staging.iterdir() if p.is_file())
                job['staged_filenames'] = names
        total = len(names)
        if total == 0:
            return False, {'error': 'No files uploaded for this job'}
        if append is not None:
            job['append'] = bool(append)
        use_append = bool(job.get('append'))
        started_at = time.time()
        job['status'] = 'started'
        job['started_at'] = started_at
        job['processed_files'] = 0
        job['failed_files'] = 0
        job['failed_filenames'] = []
        job['success_filenames'] = []
        job['results'] = []
        job['message'] = 'Processing...'
        job['total_files'] = total

    worker_id = f'pid-{os.getpid()}-{uuid.uuid4().hex[:8]}'
    claimed = claim_session_lease(job_id, worker_id, total_files=total)
    if not claimed:
        with _local_jobs_lock:
            j = _local_jobs.get(job_id)
            if j and j.get('status') == 'started':
                j['status'] = 'pending'
                j['message'] = 'Waiting for worker lease...'
        return False, {
            'error': 'Job is already being processed by another worker',
            'status': 'started',
        }

    t = threading.Thread(
        target=_worker,
        args=(job_id, started_at, use_append, worker_id),
        daemon=True,
    )
    t.start()
    return True, {
        'job_id': job_id,
        'total_files': total,
        'status': 'started',
        'message': 'Local bulk parsing started.',
    }


def start_local_job(files_list: list[tuple[str, bytes]], started_by=None, append: bool = False) -> tuple[str, dict]:
    """
    Backward-compatible: create job, stage all files, start immediately.
    Returns (job_id, payload).
    """
    job_id, _ = create_local_job(started_by=started_by, append=append)
    ok, staged = stage_files(job_id, files_list, started_by=started_by)
    if not ok:
        with _local_jobs_lock:
            _local_jobs.pop(job_id, None)
        raise ValueError(staged.get('error', 'Upload failed'))
    ok, result = start_staged_job(job_id, append=append)
    if not ok:
        raise ValueError(result.get('error', 'Start failed'))
    return job_id, result


def get_local_progress(job_id: str, check_only: bool = False) -> tuple[bool, dict]:
    """Return (True, progress_dict) if job exists; else (False, error_dict)."""
    with _local_jobs_lock:
        job = _local_jobs.get(job_id)
    if job:
        if check_only:
            return True, {'started_by': job.get('started_by')}
        return True, {
            'status': job['status'],
            'started_by': job.get('started_by'),
            'total_files': job['total_files'],
            'processed_files': job.get('processed_files', 0),
            'failed_files': job.get('failed_files', 0),
            'message': job.get('message', ''),
            'failed_filenames': job.get('failed_filenames', []),
            'success_filenames': job.get('success_filenames', []),
            'queued_filenames': [
                name
                for name in (job.get('staged_filenames') or [])
                if name
                and name not in (job.get('success_filenames') or [])
                and name not in (job.get('failed_filenames') or [])
            ],
            'failed_details': job.get('failed_details', []),
        }

    from app.domains.administration.repositories.bulk_session_db import get_session_owner, get_session_progress

    db_progress = get_session_progress(job_id)
    if db_progress:
        if check_only:
            return True, {'started_by': db_progress.get('started_by')}
        return True, db_progress

    owner = get_session_owner(job_id)
    if owner and check_only:
        return True, {'started_by': owner}

    return False, {'error': 'Job not found'}


def get_local_download(job_id: str) -> tuple[bool, Any]:
    """
    If job completed, return (True, (bytes_io, filename, content_type)).
    Else (False, error_dict).
    """
    export_file = _export_path(job_id)
    with _local_jobs_lock:
        job = _local_jobs.get(job_id)
    if job:
        if job['status'] != 'completed':
            return False, {'error': 'Job not completed yet'}
        if export_file.is_file():
            return True, (
                io.BytesIO(export_file.read_bytes()),
                'Parsed_Resumes.xlsx',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
        rows = job.get('results') or []
        try:
            xlsx_bytes = _build_excel_bytes(rows)
        except Exception as e:
            return False, {'error': f'Excel build failed: {e}'}
        return True, (
            io.BytesIO(xlsx_bytes),
            'Parsed_Resumes.xlsx',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    if export_file.is_file():
        return True, (
            io.BytesIO(export_file.read_bytes()),
            'Parsed_Resumes.xlsx',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    from app.domains.administration.repositories.bulk_session_db import get_session_progress

    db_progress = get_session_progress(job_id)
    if db_progress and db_progress.get('status') == 'completed':
        return False, {'error': 'Export file not found for completed session'}

    return False, {'error': 'Job not found'}
