"""Bulk Excel Form DTO flatten + Field Trace + det-skip gate."""
from __future__ import annotations

import io
import os
import sys
from pathlib import Path

import pytest

BACKEND_ROOT = Path(__file__).resolve().parents[3] / 'apps' / 'backend'
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

os.environ.setdefault('RESUME_SKIP_LLM_WHEN_DETERMINISTIC', 'true')
os.environ.setdefault('DOCUMENT_INTELLIGENCE_SEMANTIC_AI', 'false')

from app.ai.document_intelligence.models.form_dtos import (
    ApplicationFormDTO,
    EducationFormRow,
    ExperienceFormRow,
    FieldTrace,
)
from app.ai.parser.enrichment.resume_text_inference import is_plausible_location_value
from app.workers.bulk_parser import (
    _apply_coverage_parse_honesty,
    _build_excel_bytes,
    _bulk_det_skip_ok,
    _bulk_needs_ocr_retry,
    _failed_excel_row,
    _flatten_toon,
    _reconcile_missing_excel_rows,
)


def _form(**kwargs) -> ApplicationFormDTO:
    return ApplicationFormDTO(**kwargs)


def test_flatten_prefers_form_dto_over_toon():
    toon = {
        'person': {
            'name': 'TOON NAME',
            'email': 'toon@example.com',
            'phone': '9999999999',
            'location': 'TOON City',
        },
        'skills': ['TOONSKILL'],
        'experience': [{'title': 'Junk Title', 'company': 'JunkCo', 'from': '2020-01'}],
        'education': [{'degree': 'Fake', 'institution': 'Nowhere'}],
        'summary': 'toon summary',
        'total_experience_years': None,
    }
    form = _form(
        fullName='Form Name',
        email='form@example.com',
        phone='9876543210',
        currentLocation='Pune',
        preferredLocation='Pune',
        skills='Python, SQL',
        skillsList=['Python', 'SQL'],
        summary='form summary',
        education=[EducationFormRow(degree='B.Tech', institution='Pune University', endMonth='2024')],
        experiences=[
            ExperienceFormRow(
                company='Acme',
                role='Intern',
                startMonth='2024-01',
                endMonth='2024-06',
                description='Worked Jan 2024 - Jun 2024 on APIs.',
            )
        ],
        coverage=[
            {'field': 'fullName', 'status': 'filled', 'evidence': True},
            {'field': 'email', 'status': 'filled', 'evidence': True},
            {'field': 'phone', 'status': 'filled', 'evidence': True},
            {'field': 'location', 'status': 'filled', 'evidence': True},
            {'field': 'education', 'status': 'filled', 'evidence': True},
            {'field': 'experience', 'status': 'filled', 'evidence': True},
        ],
        trace=[
            FieldTrace(
                form_field='preferredLocation',
                canonical_path='contact.location',
                mapper='test',
                reason='fallback_current_location',
            )
        ],
    )
    text = (
        'Form Name\nform@example.com | 9876543210 | Pune\n'
        'SKILLS\nPython, SQL\n'
        'EXPERIENCE\nIntern at Acme (2024-01-2024-06)\n'
        'EDUCATION\nB.Tech - Pune University [2024]\n'
    )
    row = _flatten_toon(toon, 'x.pdf', form=form, raw_text=text)
    assert row['Name'] == 'Form Name'
    assert row['Email'] == 'form@example.com'
    assert 'Python' in row['Skills']
    assert 'TOONSKILL' not in row['Skills']
    assert 'Intern at Acme' in row['Experience']
    assert 'Junk' not in row['Experience']
    assert 'B.Tech' in row['Education']
    assert row['Total Experience Years'] not in ('', None)
    assert float(row['Total Experience Years']) > 0
    traces = {t['Field']: t for t in row['_field_trace']}
    assert traces['preferred_location']['Verdict'] == 'fallback'
    assert traces['name']['Verdict'] == 'ok'


def test_field_trace_weak_missing_when_evidence_in_resume():
    text = """
Priya Sharma
priya@example.com | 9876543210 | Mumbai
SKILLS
Python
EXPERIENCE
Software Intern - ValueDX - Jun 2023 - Aug 2023
EDUCATION
B.Tech - Mumbai University
"""
    form = _form(
        fullName='Priya Sharma',
        email='priya@example.com',
        phone='9876543210',
        currentLocation='Mumbai',
        skillsList=['Python'],
        coverage=[
            {'field': 'fullName', 'status': 'filled', 'evidence': True},
            {'field': 'email', 'status': 'filled', 'evidence': True},
            {'field': 'phone', 'status': 'filled', 'evidence': True},
            {'field': 'location', 'status': 'filled', 'evidence': True},
            {'field': 'education', 'status': 'filled', 'evidence': True},
            {'field': 'experience', 'status': 'missing_with_evidence', 'evidence': True},
        ],
    )
    toon = {
        'person': {'name': 'Priya Sharma', 'email': 'priya@example.com', 'phone': '9876543210'},
        'skills': ['Python'],
        'experience': [],
        'education': [{'degree': 'B.Tech', 'institution': 'Mumbai University'}],
    }
    row = _flatten_toon(toon, 'priya.pdf', form=form, raw_text=text)
    traces = {t['Field']: t for t in row['_field_trace']}
    assert traces['experience']['Verdict'] == 'weak_missing'
    status = _apply_coverage_parse_honesty(
        row, form, parse_status='ok', note_bits=['source=engine:deterministic']
    )
    assert status == 'partial'
    assert 'trace_weak=' in row['ParseNotes']
    assert 'experience' in row['ParseNotes']


def test_field_trace_weak_ungrounded_location():
    form = _form(
        fullName='Aarya',
        email='a@example.com',
        phone='9876543210',
        currentLocation='HTML, JS',
        coverage=[
            {'field': 'fullName', 'status': 'filled', 'evidence': True},
            {'field': 'email', 'status': 'filled', 'evidence': True},
            {'field': 'location', 'status': 'filled', 'evidence': True},
        ],
    )
    toon = {'person': {'name': 'Aarya', 'email': 'a@example.com', 'phone': '9876543210'}}
    text = 'Aarya\na@example.com\nPune\nSKILLS\nHTML, JS\n'
    row = _flatten_toon(toon, 'aarya.pdf', form=form, raw_text=text)
    traces = {t['Field']: t for t in row['_field_trace']}
    assert traces['location']['Verdict'] == 'weak_ungrounded'


def test_field_trace_fresher_experience_absent():
    text = """
Test User
t@example.com | 9876543210 | Pune
SKILLS
Python, SQL
EDUCATION
B.Tech - Test University
"""
    form = _form(
        fullName='Test User',
        email='t@example.com',
        phone='9876543210',
        currentLocation='Pune',
        skillsList=['Python', 'SQL'],
        education=[EducationFormRow(degree='B.Tech', institution='Test University')],
        coverage=[
            {'field': 'fullName', 'status': 'filled', 'evidence': True},
            {'field': 'email', 'status': 'filled', 'evidence': True},
            {'field': 'phone', 'status': 'filled', 'evidence': True},
            {'field': 'location', 'status': 'filled', 'evidence': True},
            {'field': 'education', 'status': 'filled', 'evidence': True},
            {'field': 'experience', 'status': 'missing_no_evidence', 'evidence': False},
        ],
    )
    toon = {
        'person': {'name': 'Test User', 'email': 't@example.com', 'phone': '9876543210'},
        'skills': ['Python'],
        'experience': [],
        'education': [{'degree': 'B.Tech', 'institution': 'Test University'}],
    }
    row = _flatten_toon(toon, 'fresher.pdf', form=form, raw_text=text)
    traces = {t['Field']: t for t in row['_field_trace']}
    assert traces['experience']['Verdict'] == 'absent'


def test_excel_bytes_include_field_trace_sheet():
    form = _form(fullName='A', email='a@a.com', phone='9876543210')
    row = _flatten_toon(
        {'person': {'name': 'A', 'email': 'a@a.com', 'phone': '9876543210'}},
        'a.pdf',
        form=form,
        raw_text='A a@a.com 9876543210',
    )
    data = _build_excel_bytes([row])
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True)
    assert 'Resumes' in wb.sheetnames
    assert 'Field Trace' in wb.sheetnames
    wb.close()


def test_det_skip_refused_when_location_missing_with_evidence():
    toon = {
        'type': 'resume',
        'person': {
            'name': 'Priya Sharma',
            'email': 'priya.sharma@example.com',
            'phone': '9876543210',
        },
        'skills': ['Python', 'SQL'],
        'education': [{'degree': 'B.Tech', 'institution': 'Mumbai University'}],
        'experience': [
            {'title': 'Software Intern', 'company': 'Acme', 'from': '2023-06', 'to': '2023-08'}
        ],
    }
    form = _form(
        fullName='Priya Sharma',
        email='priya.sharma@example.com',
        phone='9876543210',
        skillsList=['Python', 'SQL'],
        education=[EducationFormRow(degree='B.Tech', institution='Mumbai University')],
        experiences=[
            ExperienceFormRow(role='Software Intern', company='Acme', startMonth='2023-06', endMonth='2023-08')
        ],
        coverage=[
            {'field': 'fullName', 'status': 'filled', 'evidence': True},
            {'field': 'email', 'status': 'filled', 'evidence': True},
            {'field': 'phone', 'status': 'filled', 'evidence': True},
            {'field': 'location', 'status': 'missing_with_evidence', 'evidence': True},
            {'field': 'education', 'status': 'filled', 'evidence': True},
            {'field': 'experience', 'status': 'filled', 'evidence': True},
        ],
    )
    text = 'Priya Sharma\npriya.sharma@example.com\nMumbai\nSoftware Intern at Acme\nB.Tech Mumbai University'
    row = _flatten_toon(toon, 'p.pdf', form=form, raw_text=text)
    assert _bulk_det_skip_ok(
        toon=toon,
        form=form,
        raw_text=text,
        field_trace=row.get('_field_trace') or [],
    ) is False


def test_location_rejects_cv_title_and_skill_pairs():
    assert not is_plausible_location_value('Curriculum Vitae')
    assert not is_plausible_location_value('Patching, Ansible')
    assert not is_plausible_location_value('Business Communication, Financial')
    assert not is_plausible_location_value('Oracle DBA')
    assert not is_plausible_location_value('Database Administrator')
    assert is_plausible_location_value('Pune')
    assert is_plausible_location_value('Kalwa, Maharashtra')


def test_duty_fragments_not_job_headers():
    from app.ai.document_intelligence.parsers.resume import _is_bullet_or_duty_line
    from app.ai.parser.enrichment.resume_text_inference import is_plausible_job_title

    assert _is_bullet_or_duty_line('for multiple services.')
    assert _is_bullet_or_duty_line('and visualization')
    assert not is_plausible_job_title('for multiple services.')


def test_education_table_pipe_row():
    from app.ai.document_intelligence.parsers.resume import parse_education

    rows = parse_education(
        'Degree | Institution | Year\nB.Tech | Pune University | 2024\n'
    )
    assert any(
        'B.Tech' in (e.degree or '') and 'Pune' in (e.institution or '') for e in rows
    )


def test_flatten_years_from_current_job_without_end():
    form = _form(
        fullName='A',
        email='a@example.com',
        phone='9876543210',
        experiences=[
            ExperienceFormRow(
                company='Acme',
                role='Engineer',
                startMonth='2023-01',
                endMonth='',
                isCurrent=True,
            )
        ],
    )
    toon = {
        'person': {'name': 'A', 'email': 'a@example.com', 'phone': '9876543210'},
        'total_experience_years': '',
        'experience': [{'title': 'Engineer', 'company': 'Acme'}],
    }
    text = 'A\na@example.com\nEngineer at Acme Jan 2023 - Present'
    row = _flatten_toon(toon, 'a.pdf', form=form, raw_text=text)
    assert row['Total Experience Years'] not in ('', None)
    assert float(row['Total Experience Years']) >= 1.0


def test_years_absent_when_only_graduation_years():
    text = """
Test User
t@example.com | 9876543210 | Pune
SKILLS
Python
EDUCATION
B.Tech - Test University, 2024
"""
    form = _form(
        fullName='Test User',
        email='t@example.com',
        phone='9876543210',
        currentLocation='Pune',
        skillsList=['Python'],
        education=[EducationFormRow(degree='B.Tech', institution='Test University', endMonth='2024')],
        coverage=[
            {'field': 'fullName', 'status': 'filled', 'evidence': True},
            {'field': 'education', 'status': 'filled', 'evidence': True},
            {'field': 'experience', 'status': 'missing_no_evidence', 'evidence': False},
        ],
    )
    toon = {
        'person': {'name': 'Test User', 'email': 't@example.com', 'phone': '9876543210'},
        'education': [{'degree': 'B.Tech', 'institution': 'Test University'}],
    }
    row = _flatten_toon(toon, 'grad.pdf', form=form, raw_text=text)
    traces = {t['Field']: t for t in row['_field_trace']}
    assert traces['years']['Verdict'] == 'absent'


def test_unlabeled_education_footer_header():
    from app.ai.document_intelligence.parsers.resume import parse_education

    text = """
DEVIN D'SILVA
t@example.com

Bachelor of Science in Information Technology
University of Mumbai | 2023 – 2026
Grade : 8.93/10
Higher Secondary Education (Commerce)
Maharashtra State Board | 2021 – 2023

Certifications
Education
Skills
"""
    rows = parse_education('Education', text)
    assert any(
        'Bachelor' in (e.degree or '') and 'Mumbai' in (e.institution or '') for e in rows
    )


def test_prefix_sidebar_dates_attach_to_undated_jobs():
    from app.ai.document_intelligence.parsers.resume import parse_experience

    full = """
31/01/2026 - 30/06/2026
2018
2020
DHAVAL RANE
Mumbai

Experience
ELV Link Technologies Private Limited
Graduate Engineer Trainee
Worked at ELV Link Technologies Private Limited as a Graduate Engineer Trainee.
"""
    section = """
ELV Link Technologies Private Limited
Graduate Engineer Trainee
Worked at ELV Link Technologies Private Limited as a Graduate Engineer Trainee.
"""
    rows = parse_experience(section, full)
    assert rows
    assert any((e.start or '') for e in rows)


def test_unlabeled_dated_remote_job_appears_in_excel():
    form = _form(
        fullName='Sandeep',
        email='s@example.com',
        phone='9876543210',
        experiences=[
            ExperienceFormRow(
                company='',
                role='',
                startMonth='2025-07',
                endMonth='2025-10',
                description='Designed and managed CI/CD pipelines with GitHub Actions.',
            )
        ],
    )
    toon = {'person': {'name': 'Sandeep', 'email': 's@example.com', 'phone': '9876543210'}}
    text = (
        'Sandeep\ns@example.com\nExperience\n07/2025 – 10/2025 | Remote\n'
        'Designed and managed CI/CD pipelines with GitHub Actions.'
    )
    row = _flatten_toon(toon, 'jorka.pdf', form=form, raw_text=text)
    assert 'CI/CD' in (row['Experience'] or '')
    assert row['Total Experience Years'] not in ('', None)


def test_failed_excel_row_and_reconcile_match_file_count():
    ok = _flatten_toon(
        {'person': {'name': 'A', 'email': 'a@a.com', 'phone': '9876543210'}},
        'ok.pdf',
        raw_text='A a@a.com 9876543210',
    )
    fail = _failed_excel_row('bad.pdf', code='insufficient_text', message='too short')
    assert fail['ParseStatus'] == 'failed'
    assert fail['Filename'] == 'bad.pdf'
    mixed = _reconcile_missing_excel_rows([ok, fail], ['ok.pdf', 'bad.pdf', 'missing.pdf'])
    assert len(mixed) == 3
    by_name = {r['Filename']: r for r in mixed}
    assert by_name['missing.pdf']['ParseStatus'] == 'failed'
    assert 'not_processed' in by_name['missing.pdf']['ParseNotes']
    data = _build_excel_bytes(mixed)
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True)
    ws = wb['Resumes']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    # header + 3 data rows
    assert len(rows) == 4
    statuses = {r[0]: r[-2] for r in rows[1:]}
    assert statuses['bad.pdf'] == 'failed'
    assert statuses['missing.pdf'] == 'failed'


def test_ocr_retry_on_extract_err_even_when_text_present():
    def never_garbage(_s):
        return False

    assert _bulk_needs_ocr_retry('pdf', 'plenty of extracted text here ' * 20, 'boom', looks_like_garbage=never_garbage)
    assert _bulk_needs_ocr_retry('pdf', '', None, looks_like_garbage=never_garbage)
    assert not _bulk_needs_ocr_retry(
        'docx', 'plenty of extracted text here ' * 20, 'boom', looks_like_garbage=never_garbage
    )
    assert not _bulk_needs_ocr_retry(
        'pdf', 'plenty of extracted text here ' * 20, None, looks_like_garbage=never_garbage
    )


def test_company_emdash_role_dates():
    from app.ai.document_intelligence.parsers.resume import parse_experience

    rows = parse_experience(
        'ELV Link Technologies Pvt. Ltd. — Graduate Trainee Engineer (Jan 2026 – Present)\n'
        'Assisting in installation and maintenance.\n'
    )
    assert any(
        'ELV' in (e.company or '') and 'Trainee' in (e.role or '') and (e.start or '')
        for e in rows
    )
    assert any(e.is_current for e in rows)


def test_company_city_pipe_role_dates():
    from app.ai.document_intelligence.parsers.resume import parse_experience

    rows = parse_experience(
        'Realatte Ventures Limited, Andheri (E) | Full Stack Developer, April 2026 – July 2026\n'
        'Developed Shopify apps.\n'
    )
    assert any(
        'Realatte' in (e.company or '') and 'Full Stack' in (e.role or '')
        for e in rows
    )


def test_nashik_alias_is_grounded_against_nasik_source():
    form = _form(
        fullName='Saloni',
        email='s@example.com',
        phone='8080234411',
        currentLocation='Nashik',
        coverage=[{'field': 'location', 'status': 'filled', 'evidence': True}],
    )
    toon = {'person': {'name': 'Saloni', 'email': 's@example.com', 'phone': '8080234411'}}
    text = 'Saloni Khivansara\nNasik, India\ns@example.com\n'
    row = _flatten_toon(toon, 'saloni.pdf', form=form, raw_text=text)
    traces = {t['Field']: t for t in row['_field_trace']}
    assert traces['location']['Verdict'] == 'ok'


def test_qualification_dingbat_and_college_first_education():
    from app.ai.document_intelligence.parsers.resume import parse_education

    dingbat = parse_education(
        'QUALIFICATION :\n Bachelor of Computer Application (BCA)\n'
        'Mahatma Gandhi Kashi Vidyapith University in Varanasi\n'
        'Session- July 2022 - July 2025\n',
        '',
    )
    assert any(
        'Bachelor' in (e.degree or '') and 'Vidyapith' in (e.institution or '')
        for e in dingbat
    )
    body = parse_education(
        '',
        'Jai Hind College, Mumbai\nMaster\'s of Science\n'
        'Graduated with high honors with a Master of Science in Big Data Analytics.\n'
        'PRAPTI DALVI\nprapti@example.com\n',
    )
    assert any(
        'Master' in (e.degree or '') and 'Jai Hind' in (e.institution or '')
        for e in body
    )


def test_workexperience_glued_header_is_experience():
    from app.ai.parser.layout.heuristic import normalize_section_header

    assert normalize_section_header('WORKEXPERIENCE') == 'Experience'
    assert normalize_section_header('QUALIFICATION :') == 'Education'


