"""
Local bulk resume parsing when external Bulk-Resume-Parser (port 8001) is unavailable.
Uses existing text_extraction + llm_service; stores results in memory and exports Excel.
"""
import io
import threading
import uuid
from typing import Any

# In-memory job store: job_id -> { status, total_files, processed_files, failed_files, results, message }
_local_jobs: dict[str, dict[str, Any]] = {}
_local_jobs_lock = threading.Lock()

ALLOWED_EXT = {'pdf', 'doc', 'docx'}


def _flatten_toon(toon: dict, filename: str) -> dict:
    """Flatten one TOON resume to a row dict for Excel."""
    person = toon.get('person') or {}
    skills = toon.get('skills') or []
    exp = toon.get('experience') or []
    edu = toon.get('education') or []
    certs = toon.get('certifications') or []
    exp_text = '; '.join(
        f"{e.get('title', '')} at {e.get('company', '')} ({e.get('from', '')}-{e.get('to', '')})"
        for e in exp[:10]
    )
    edu_text = '; '.join(
        f"{e.get('degree', '')} - {e.get('institution', '')}"
        for e in edu[:5]
    )
    return {
        'Filename': filename,
        'Name': person.get('name') or '',
        'Email': person.get('email') or '',
        'Phone': person.get('phone') or '',
        'LinkedIn': person.get('linkedin') or '',
        'GitHub': person.get('github') or '',
        'Summary': (toon.get('summary') or '')[:500],
        'Skills': ', '.join(skills[:30]) if isinstance(skills, list) else str(skills)[:500],
        'Experience': exp_text[:1000],
        'Education': edu_text[:500],
        'Certifications': ', '.join(certs[:20]) if isinstance(certs, list) else str(certs)[:300],
        'Total Experience Years': toon.get('total_experience_years') or '',
    }


def _build_excel_bytes(rows: list[dict]) -> bytes:
    """Build .xlsx from list of row dicts using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumes"
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, '') for h in headers])
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = min(40, max(10, len(str(headers[col - 1])) + 2))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _worker(job_id: str, files_list: list[tuple[str, bytes]]) -> None:
    """Background: extract text + LLM per file, append to job results."""
    from text_extraction import extract_text
    from llm_service import call_llm

    with _local_jobs_lock:
        job = _local_jobs.get(job_id)
    if not job or job.get('status') == 'cancelled':
        return
    total = len(files_list)
    results = []
    processed = 0
    failed = 0
    for filename, data in files_list:
        with _local_jobs_lock:
            j = _local_jobs.get(job_id)
        if not j or j.get('status') == 'cancelled':
            break
        try:
            raw_text = extract_text(data, filename)
            if not raw_text or len(raw_text.strip()) < 30:
                failed += 1
                with _local_jobs_lock:
                    _local_jobs[job_id]['processed_files'] = processed + failed
                    _local_jobs[job_id]['failed_files'] = failed
                    _local_jobs[job_id]['message'] = f'Skipped (insufficient text): {filename}'
                continue
            toon = call_llm(raw_text, 'resume')
            row = _flatten_toon(toon, filename)
            results.append(row)
            processed += 1
        except Exception as e:
            failed += 1
            with _local_jobs_lock:
                _local_jobs[job_id]['message'] = f'Failed: {filename} - {str(e)[:100]}'
        with _local_jobs_lock:
            _local_jobs[job_id]['processed_files'] = processed + failed
            _local_jobs[job_id]['failed_files'] = failed
    with _local_jobs_lock:
        if job_id in _local_jobs:
            _local_jobs[job_id]['status'] = 'completed'
            _local_jobs[job_id]['processed_files'] = processed + failed
            _local_jobs[job_id]['failed_files'] = failed
            _local_jobs[job_id]['results'] = results
            _local_jobs[job_id]['message'] = f'Completed: {processed} successful, {failed} failed'


def start_local_job(files_list: list[tuple[str, bytes]]) -> tuple[str, dict]:
    """
    Start local bulk parse job. Returns (job_id, { job_id, total_files, status: 'started' }).
    """
    job_id = str(uuid.uuid4())
    with _local_jobs_lock:
        _local_jobs[job_id] = {
            'status': 'started',
            'total_files': len(files_list),
            'processed_files': 0,
            'failed_files': 0,
            'results': [],
            'message': 'Processing...',
        }
    t = threading.Thread(target=_worker, args=(job_id, files_list), daemon=True)
    t.start()
    return job_id, {
        'job_id': job_id,
        'total_files': len(files_list),
        'status': 'started',
        'message': 'Local bulk parsing started.',
    }


def get_local_progress(job_id: str) -> tuple[bool, dict]:
    """Return (True, progress_dict) if job exists; else (False, error_dict)."""
    with _local_jobs_lock:
        job = _local_jobs.get(job_id)
    if not job:
        return False, {'error': 'Job not found'}
    return True, {
        'status': job['status'],
        'total_files': job['total_files'],
        'processed_files': job.get('processed_files', 0),
        'failed_files': job.get('failed_files', 0),
        'message': job.get('message', ''),
    }


def get_local_download(job_id: str) -> tuple[bool, Any]:
    """
    If job completed, return (True, (bytes_io, filename, content_type)).
    Else (False, error_dict).
    """
    with _local_jobs_lock:
        job = _local_jobs.get(job_id)
    if not job:
        return False, {'error': 'Job not found'}
    if job['status'] != 'completed':
        return False, {'error': 'Job not completed yet'}
    rows = job.get('results') or []
    try:
        xlsx_bytes = _build_excel_bytes(rows)
    except Exception as e:
        return False, {'error': f'Excel build failed: {e}'}
    return True, (io.BytesIO(xlsx_bytes), 'Parsed_Resumes.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
